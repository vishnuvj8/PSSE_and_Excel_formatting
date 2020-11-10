

import os
import openpyxl
import re
os.getcwd()

name_1=open("DPP_APR18_Central_MTEP18_Master_Fault_List_FINAL_01212020_1_100.swi", "r")
content=name_1.read()

content_1 = re.sub('Description   ','Description ', content)

miso_list = content_1.split('Nomore')

last_row = len(miso_list)
fault_type = []
cycles = []
descrep = []
fault_bus = []
con_events = []
con_events_1 = []

for i in range (last_row-1):
    name_2 = miso_list[i].split(' ')
    name_rows = len(name_2)
    len(name_2)


## to get cycles and name##
print "-------------------------------------------------------Start------------------------------------------------------"

for i in range (last_row-1):
    name_2 = miso_list[i].split('\n')
    x = len(name_2)
    for j in range (x):
        name_3 = name_2[j].split(' ')
        name_4 = name_2[j].split(' ')
        name_5 = name_2[j].split(' ')
        y = len(name_3)
        z = len(name_4)                 
        for l in range (z):
            if name_4[l]=="Description":
                descrep.append (name_4[l+1])
        for k in range (y):
            if name_3[k]== "After":                
                cycles.append (name_3[k+1])
    cycles.append ("Nomore")
    descrep.append ("Nomore")


print "-------------------------------------------------------15%------------------------------------------------------"

## to get fault bus number ##

for m in range (last_row-1):
    temp_1 = miso_list[m].split('\n')
    a = len (temp_1)
    for n in range (a):
        temp_2 = temp_1[n].split(' ')
        b= len(temp_2)
        for o in range (b):
            if temp_2[o]=="Three":
                if temp_2[o+1]=="phase" or "Phase":
                    if temp_2[o+2]=="fault" or "Fault":
                        if temp_2[o-1]=="clear":
                            pass
                        else:
                            fault_bus.append (temp_2[o+5])
                            fault_iden_3 = True
                            fault_iden_1 = False

            if temp_2[o]=="One":
                if temp_2[o+1]=="phase" or "Phase":
                    if temp_2[o+2]=="to" or "To":
                        if temp_2[o+3]=="ground" or "Ground":
                            if temp_2[o+4]=="fault" or "Fault":                                
                                if temp_2[o-1]=="clear":
                                    pass
                                else:
                                    fault_bus.append (temp_2[o+7])
                                    fault_iden_1 = True
                                    fault_iden_3 = False
    fault_bus.append("Nomore")
    if fault_iden_3 == True:
        fault_type.append ("3PH")
    else:
        fault_iden_1 == True
        fault_type.append("1PH")

print "-------------------------------------------------------30%------------------------------------------------------"

for p in range (last_row-1):
    temp_3 = miso_list[p].split('\n')
    c = len (temp_3)
    for q in range (c):
        temp_4 = temp_3[q].split(' ')
        d = len(temp_4)
        for r in range (d):
            if temp_4[r]=="Reduce":
                con_events.append(temp_3[q])
            if temp_4[r]=="Reconnect":
                con_events.append(temp_3[q])
            if temp_4[r]=="Shed":
                con_events.append(temp_3[q])
            if temp_4[r]=="Modify":
                if temp_4[r+1]=="Line":
                    con_events.append(temp_3[q])
            if temp_4[r]=="Add":
                if temp_4[r+1]=="Line":
                    con_events.append(temp_3[q])
            if temp_4[r]=="Add":
                if temp_4[r+1]=="line":
                    con_events.append(temp_3[q])
            if temp_4[r]== "tap":
                con_events.append(temp_3[q])
            if temp_4[r]=="Tap":
                con_events.append(temp_3[q])
            if temp_4[r] == "outage":
                con_events.append(temp_3[q])
            if temp_4[r]=="Outage":
                con_events.append(temp_3[q])
            if temp_4[r] == "Disconnect":
                con_events.append(temp_3[q])
            else:
                pass
            if temp_4[r]== "Remove":
                    if temp_4[r + 1] == "Admittance":
                        pass
                    else:
                        con_events.append(temp_3[q])
    con_events.append("Nomore")

print "-------------------------------------------------------40%------------------------------------------------------"
x=len(con_events)

for s in range (x):
    if con_events[s][0]=="/":
        pass
    else:
        con_events_1.append(con_events[s])

y=len(con_events_1)

## to change bus numbers to names ##
from openpyxl import load_workbook
wc = load_workbook('Planning_data_directory.xlsx')
sheet_1 = wc.get_sheet_by_name('Sheet1')
#print sheet_1['A15'].value

bus_num = []
bus_name=[]
for z in range(sheet_1.max_row):
    sheet_1.cell(row=z+1, column=1).value = bus_num.append(sheet_1.cell(row=z+1, column=1).value)
    sheet_1.cell(row=z + 1, column=2).value = bus_name.append(sheet_1.cell(row=z + 1, column=2).value)

bus_name_size = len (bus_name)
bus_num_size = len (bus_num)
fault_bus_size = len (fault_bus)
for f in range (fault_bus_size):
    for g in range (bus_num_size):
        if fault_bus[f]==bus_num[g]:
            print(g, f)
            fault_bus[f]= bus_name[g]

print "-------------------------------------------------------45%------------------------------------------------------"
#print fault_bus
temp_6 = []
con_events_1_size = len(con_events_1)
for h in range(con_events_1_size):
    temp_5 = con_events_1[h].split(' ')
    temp_5_size = len(temp_5)
    for q in range(temp_5_size):
        for g in range(bus_num_size):
            if temp_5[q] == bus_num[g]:
                print(g, q, h)
                temp_5[q] = bus_name[g]
        temp_6 = str(temp_6) + " " + str(temp_5[q])
    temp_6 = str(temp_6) + '\n'
#print temp_6

con_eventa_final = temp_6.split('Nomore')


temp_7 = []
temp_8= []
temp_9 = []
temp_10 =[]
temp_11 = []
temp_12 =[]
cycles_size = len(cycles)
fault_bus_size = len (fault_bus)
for dd in range (cycles_size):
    temp_7 = str(temp_7) + str(cycles[dd]) + '\n'

for ee in range(fault_bus_size):
    temp_8 = str(temp_8) + str(fault_bus[ee]) + '\n'

cycles_final = temp_7.split('Nomore')
fault_bus_final = temp_8.split('Nomore')

descrep_size = len(descrep)
for ee in range (descrep_size):
    temp_9 = str(temp_9) + str(descrep[ee] + '\n')

descrep_final  = temp_9.split('_')
descrep_final_size = len(descrep_final)
for ff in range(descrep_final_size):
    if descrep_final[ff][0] == "p" or descrep_final[ff][0] == "P" or descrep_final[ff][0] == "e" or descrep_final[ff][0] == "E":
        if descrep_final[ff][1] == "1" or descrep_final[ff][1] == "2" or descrep_final[ff][1] == "3" or descrep_final[ff][1] == "4" or descrep_final[ff][1] == "5" or descrep_final[ff][1] == "6" or descrep_final[ff][1] == "7":
            temp_10.append(descrep_final[ff])
        else:
            pass

temp_10_size = len(temp_10)
for gg in range (temp_10_size):
    temp_11 = temp_10[gg].split('\n')
    temp_12.append(temp_11[0])



print "-------------------------------------------------------90%------------------------------------------------------"
wb = openpyxl.Workbook() # Create a blank workbook.
wb.sheetnames # It starts with one sheet.
sheet = wb.active
sheet.title = 'SWI_formatted_to_excel'

wb.save("Swi_to_excel.xlsx")

wc = openpyxl.load_workbook("Swi_to_excel.xlsx")
SWI_formatted_to_excel = wb.active
SWI_formatted_to_excel.cell(row=2, column=2).value = "Reference no."
SWI_formatted_to_excel.cell(row=2, column=3).value = "Category"
SWI_formatted_to_excel.cell(row=2, column=4).value = "Contingency name"
SWI_formatted_to_excel.cell(row=2, column=5).value = "Faulted Bus"
SWI_formatted_to_excel.cell(row=2, column=6).value = "Fault"
SWI_formatted_to_excel.cell(row=2, column=7).value = "Hung Breaker"
SWI_formatted_to_excel.cell(row=2, column=8).value = "Cycles to Clear Fault"
SWI_formatted_to_excel.cell(row=2, column=9).value = "Contingency Outage"

for bb in range (last_row-1):
    SWI_formatted_to_excel.cell(row=bb+3, column=2).value = bb+1
    SWI_formatted_to_excel.cell(row=bb + 3, column=6).value = fault_type[bb]
    SWI_formatted_to_excel.cell(row=bb + 3, column=9).value = con_eventa_final[bb]
    SWI_formatted_to_excel.cell(row=bb + 3, column=5).value = fault_bus_final[bb]
    SWI_formatted_to_excel.cell(row=bb + 3, column=8).value = cycles_final[bb]
    SWI_formatted_to_excel.cell(row=bb + 3, column=3).value = temp_12[bb]

count_1 = 0
len_descrep = len(descrep)
for cc in range (len_descrep):
    if descrep[cc] == 'Nomore':
        pass
    else:
        SWI_formatted_to_excel.cell (row=count_1+3, column=4).value = descrep[cc]
        count_1 = count_1+1

wb.save(filename = "Swi_to_excel.xlsx")

print "@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@Finished@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@"
