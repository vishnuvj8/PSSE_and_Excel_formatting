
import os
import openpyxl
from openpyxl import load_workbook
os.getcwd()

wb = openpyxl.Workbook() # Create a blank workbook.
wb.sheetnames # It starts with one sheet.
sheet = wb.active
sheet.title = 'Excel_comparison'

wb.save("Excel_comparison.xlsx")

wc = openpyxl.load_workbook("Excel_comparison.xlsx")
Excel_comparison = wb.active
Excel_comparison.cell(row=2, column=2).value = "Contingency"
Excel_comparison.cell(row=2, column=3).value = "From_bus"
Excel_comparison.cell(row=2, column=4).value = "TO_bus"
Excel_comparison.cell(row=2, column=5).value = "ID"
Excel_comparison.cell(row=2, column=6).value = "Pre-Case-Loading"
Excel_comparison.cell(row=2, column=7).value = "Post-Case_loading"
Excel_comparison.cell(row=2, column=8).value = "Pre-Convergence?"
Excel_comparison.cell(row=2, column=9).value = "post-Convergence?"

from openpyxl import load_workbook
wc = load_workbook('pre.xlsx')
sheet_1 = wc.get_sheet_by_name('Pre')
#print sheet_1['A15'].value
print sheet_1.cell(row=74, column=1).value
print sheet_1.max_row
print sheet_1.max_column



wd = load_workbook('Post.xlsx')
sheet_2 = wd.get_sheet_by_name('Post')
#print sheet_2['A15'].value
print sheet_2.cell(row=17, column=3).value
print sheet_2.max_row
print sheet_2.max_column

contin_pre= []
from_bus_pre=[]
to_bus_pre=[]
ID_pre = []
loading_pre=[]
convergence_pre = []

contin_post= []
from_bus_post=[]
to_bus_post=[]
ID_post = []
loading_post=[]
convergence_post = []

for z in range(sheet_1.max_row):
    sheet_1.cell(row=z+1, column=1).value = contin_pre.append(sheet_1.cell(row=z+1, column=1).value)
    sheet_1.cell(row=z + 1, column=2).value = from_bus_pre.append(sheet_1.cell(row=z + 1, column=2).value)
    sheet_1.cell(row=z + 1, column=3).value = to_bus_pre.append(sheet_1.cell(row=z + 1, column=3).value)
    sheet_1.cell(row=z + 1, column=4).value = ID_pre.append(sheet_1.cell(row=z + 1, column=4).value)
    sheet_1.cell(row=z + 1, column=5).value = loading_pre.append(sheet_1.cell(row=z + 1, column=5).value)
    sheet_1.cell(row=z + 1, column=6).value = convergence_pre.append(sheet_1.cell(row=z + 1, column=6).value)

for x in range(sheet_2.max_row):
    sheet_2.cell(row=x+1, column=1).value = contin_post.append(sheet_2.cell(row=x+1, column=1).value)
    sheet_2.cell(row=x + 1, column=2).value = from_bus_post.append(sheet_2.cell(row=x + 1, column=2).value)
    sheet_2.cell(row=x + 1, column=3).value = to_bus_post.append(sheet_2.cell(row=x+ 1, column=3).value)
    sheet_2.cell(row=x + 1, column=4).value = ID_post.append(sheet_2.cell(row=x + 1, column=4).value)
    sheet_2.cell(row=x + 1, column=5).value = loading_post.append(sheet_2.cell(row=x + 1, column=5).value)
    sheet_2.cell(row=x + 1, column=6).value = convergence_post.append(sheet_2.cell(row=x + 1, column=6).value)

a=[]
b=[]
c=[]
d=[]
e=[]
f=[]
g=[]
h=[]


for i in range(sheet_2.max_row):
    for l in range(sheet_1.max_row):
            if contin_pre[l] == contin_post[i]:
                if from_bus_pre[l] == from_bus_post[i]:
                    if to_bus_pre[l] == to_bus_post[i]:
                        if ID_pre[l] == ID_post[i]:
                            a.append(contin_pre[l])
                            b.append(from_bus_pre[l])
                            c.append(to_bus_pre[l])
                            d.append(ID_pre[l])
                            e.append(loading_pre[l])
                            f.append(loading_post[i])
                            g.append(convergence_pre[l])
                            h.append(convergence_post[i])
                            #print (a)
                            print (i, l)


aa = len (a)
for bb in range (aa):
    Excel_comparison.cell(row=bb+3, column=2).value = a[bb]
    Excel_comparison.cell(row=bb+3, column=3).value = b[bb]
    Excel_comparison.cell(row=bb+3, column=4).value = c[bb]
    Excel_comparison.cell(row=bb+3, column=5).value = d[bb]
    Excel_comparison.cell(row=bb+3, column=6).value = e[bb]
    Excel_comparison.cell(row=bb+3, column=7).value = f[bb]
    Excel_comparison.cell(row=bb+3, column=8).value = g[bb]
    Excel_comparison.cell(row=bb+3, column=9).value = h[bb]


print a
print b
print c
print d
print e

wb.save("Excel_comparison.xlsx")

print "@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@Finished@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@"



